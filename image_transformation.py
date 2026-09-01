#!/usr/bin/env python3

"""
image_transformation.py

Bulk product image transformation pipeline.

Expected Excel format:

    Parent Group Id | Image Url 1 | Image Url 2 | ... | Image Url 8

Processing rules:

1. Minimum image resolution:
       Width  >= 660 px
       Height >= 900 px

2. If the product/foreground touches ANY image boundary:
       DO NOT TRANSFORM
       Status = SKIPPED_PRODUCT_TOUCHES_EDGE

3. Otherwise:
       Transform to target aspect ratio 660:900
       WITHOUT cropping or stretching.

4. Output filenames:
       ParentSKU_1.jpg
       ParentSKU_2.jpg
       ...
       ParentSKU_8.jpg

Supports:
    - Normal image URLs
    - Google Drive share URLs
    - Up to 8 images per Parent Group Id
    - Excel processing report
    - Streamlit progress callback

Dependencies:

    python3 -m pip install pillow numpy requests openpyxl
"""

import os
import re
import sys
import time
import argparse

from io import BytesIO
from collections import Counter

import numpy as np
import requests

from PIL import Image, ImageFilter, ImageOps

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

MIN_W = 660
MIN_H = 900

# H:W
TARGET_RATIO = MIN_H / MIN_W

# W:H
TARGET_W_H_RATIO = MIN_W / MIN_H

MAX_IMAGES_PER_SKU = 8

DEFAULT_SKU_COLUMN = "Parent Group Id"

RESULT_SHEET_NAME = "Processing Results"


# ============================================================
# PADDING CONFIGURATION
# ============================================================

EDGE_SAMPLE_PX = 6

SEAM_BLUR_BAND = 14
SEAM_BLUR_RADIUS = 6


# ============================================================
# DOWNLOAD CONFIGURATION
# ============================================================

MAX_DOWNLOAD_MB = 30

CONNECT_TIMEOUT = 10
READ_TIMEOUT = 30

DOWNLOAD_RETRIES = 3


# ============================================================
# EDGE DETECTION CONFIGURATION
# ============================================================

# RGB distance from estimated background after which a pixel
# is considered likely foreground/product.
BACKGROUND_DISTANCE_THRESHOLD = 35

# If more than 8% of ANY boundary differs from the background,
# treat the product as touching that edge and SKIP transformation.
EDGE_FOREGROUND_FRACTION_THRESHOLD = 0.08


# ============================================================
# IMAGE NORMALIZATION
# ============================================================


def normalize_image(img):
    """
    Correct EXIF orientation and convert image to RGB.

    Transparent images are placed onto a white background.
    """

    img = ImageOps.exif_transpose(img)

    if img.mode in ("RGBA", "LA"):

        rgba = img.convert("RGBA")

        background = Image.new(
            "RGBA",
            rgba.size,
            (255, 255, 255, 255),
        )

        background.alpha_composite(
            rgba
        )

        return background.convert(
            "RGB"
        )

    if (
        img.mode == "P"
        and "transparency" in img.info
    ):

        rgba = img.convert(
            "RGBA"
        )

        background = Image.new(
            "RGBA",
            rgba.size,
            (255, 255, 255, 255),
        )

        background.alpha_composite(
            rgba
        )

        return background.convert(
            "RGB"
        )

    return img.convert(
        "RGB"
    )


# ============================================================
# BACKGROUND DETECTION
# ============================================================


def estimate_background_color(arr):
    """
    Estimate image background using the four image corners.

    Rather than averaging all corners, choose the corner whose
    median colour is most similar to the other corners.

    This helps if the product touches one particular corner.
    """

    height, width, _ = arr.shape

    patch_size = max(
        12,
        int(
            min(
                height,
                width,
            )
            * 0.04
        ),
    )

    patch_size = min(
        patch_size,
        100,
    )

    corners = [
        arr[
            :patch_size,
            :patch_size,
            :
        ],
        arr[
            :patch_size,
            -patch_size:,
            :
        ],
        arr[
            -patch_size:,
            :patch_size,
            :
        ],
        arr[
            -patch_size:,
            -patch_size:,
            :
        ],
    ]

    corner_colors = np.asarray(
        [
            np.median(
                corner.reshape(
                    -1,
                    3,
                ),
                axis=0,
            )
            for corner in corners
        ],
        dtype=np.float32,
    )

    distances = np.linalg.norm(
        corner_colors[:, None, :]
        - corner_colors[None, :, :],
        axis=2,
    )

    medoid_index = int(
        np.argmin(
            distances.sum(
                axis=1
            )
        )
    )

    return corner_colors[
        medoid_index
    ]


def get_edge_foreground_fraction(
    edge_strip,
    background_color,
):
    """
    Estimate how much of an edge differs from the estimated
    background colour.

    Returns a value between 0 and 1.
    """

    distance = np.linalg.norm(
        edge_strip.astype(
            np.float32
        )
        - background_color.reshape(
            1,
            1,
            3,
        ),
        axis=2,
    )

    foreground_mask = (
        distance
        > BACKGROUND_DISTANCE_THRESHOLD
    )

    return float(
        foreground_mask.mean()
    )


def detect_product_touching_edges(img):
    """
    Detect whether product/foreground reaches any image boundary.

    Checks:
        TOP
        BOTTOM
        LEFT
        RIGHT

    Returns:

        warnings
        fractions

    Example:

        warnings = [
            "PRODUCT_TOUCHES_BOTTOM_EDGE",
            "PRODUCT_TOUCHES_RIGHT_EDGE"
        ]

        fractions = {
            "TOP": 0.00,
            "BOTTOM": 0.47,
            "LEFT": 0.00,
            "RIGHT": 0.28
        }
    """

    img = normalize_image(
        img
    )

    arr = np.asarray(
        img,
        dtype=np.float32,
    )

    height, width, _ = (
        arr.shape
    )

    edge_size = min(
        EDGE_SAMPLE_PX,
        height,
        width,
    )

    background_color = (
        estimate_background_color(
            arr
        )
    )

    strips = {
        "TOP": arr[
            :edge_size,
            :,
            :
        ],

        "BOTTOM": arr[
            -edge_size:,
            :,
            :
        ],

        "LEFT": arr[
            :,
            :edge_size,
            :
        ],

        "RIGHT": arr[
            :,
            -edge_size:,
            :
        ],
    }

    fractions = {
        name: get_edge_foreground_fraction(
            strip,
            background_color,
        )
        for name, strip
        in strips.items()
    }

    warnings = [
        f"PRODUCT_TOUCHES_{name}_EDGE"
        for name, fraction
        in fractions.items()
        if (
            fraction
            > EDGE_FOREGROUND_FRACTION_THRESHOLD
        )
    ]

    return (
        warnings,
        fractions,
    )


# ============================================================
# PADDING HELPERS
# ============================================================


def estimate_local_noise(strip):
    """
    Estimate local image grain/noise.
    """

    strip = strip.astype(
        np.float32
    )

    differences = []

    if strip.shape[0] > 1:

        differences.append(
            np.diff(
                strip,
                axis=0,
            ).ravel()
        )

    if strip.shape[1] > 1:

        differences.append(
            np.diff(
                strip,
                axis=1,
            ).ravel()
        )

    if not differences:

        return 0.8

    values = np.concatenate(
        differences
    )

    std = (
        values.std()
        / np.sqrt(2)
    )

    return float(
        np.clip(
            std,
            0.4,
            3.0,
        )
    )


def build_edge_fill(
    edge_strip,
    pad_size,
    axis,
):
    """
    Extend clean image background outward.

    axis="v"
        top / bottom

    axis="h"
        left / right
    """

    if pad_size <= 0:

        raise ValueError(
            "pad_size must be greater than zero"
        )

    if axis == "v":

        profile = edge_strip.mean(
            axis=0,
            keepdims=True,
        )

        return np.repeat(
            profile,
            pad_size,
            axis=0,
        )

    if axis == "h":

        profile = edge_strip.mean(
            axis=1,
            keepdims=True,
        )

        return np.repeat(
            profile,
            pad_size,
            axis=1,
        )

    raise ValueError(
        "axis must be 'v' or 'h'"
    )


def add_matched_noise(
    region,
    noise_std,
    seed,
):
    """
    Add subtle grain to generated padding.
    """

    rng = (
        np.random.default_rng(
            seed
        )
    )

    noise = rng.normal(
        0,
        noise_std,
        size=region.shape,
    )

    return np.clip(
        region.astype(
            np.float32
        )
        + noise,
        0,
        255,
    )


# ============================================================
# RESOLUTION VALIDATION
# ============================================================


def validate_resolution(
    width,
    height,
):
    """
    Validate platform minimum resolution.
    """

    if (
        width < MIN_W
        or height < MIN_H
    ):

        raise ValueError(
            f"Image "
            f"{width}x{height} "
            f"is below minimum "
            f"supported resolution "
            f"{MIN_W}x{MIN_H}"
        )


# ============================================================
# IMAGE TRANSFORMATION
# ============================================================


def pad_to_target_ratio(img):
    """
    Transform to target H:W = 900:660.

    IMPORTANT:

    This function should ONLY receive images that already passed
    product-edge detection.

    Therefore all boundaries are expected to contain background.

    No cropping.
    No stretching.
    No product distortion.

    Returns:

        transformed_image
        transformation_description
    """

    img = normalize_image(
        img
    )

    width, height = (
        img.size
    )

    validate_resolution(
        width,
        height,
    )

    current_ratio = (
        height / width
    )

    # --------------------------------------------------------
    # ALREADY AT TARGET
    # --------------------------------------------------------

    if (
        abs(
            current_ratio
            - TARGET_RATIO
        )
        < 0.001
    ):

        return (
            img,
            "NONE",
        )

    arr = np.asarray(
        img,
        dtype=np.float32,
    )

    # ========================================================
    # IMAGE TOO WIDE / SHORT
    #
    # Add TOP + BOTTOM
    # ========================================================

    if current_ratio < TARGET_RATIO:

        new_width = width

        new_height = round(
            width
            * TARGET_RATIO
        )

        pad_total = (
            new_height
            - height
        )

        pad_top = (
            pad_total
            // 2
        )

        pad_bottom = (
            pad_total
            - pad_top
        )

        canvas = np.zeros(
            (
                new_height,
                new_width,
                3,
            ),
            dtype=np.float32,
        )

        # Original image remains untouched.
        canvas[
            pad_top:
            pad_top + height,
            :,
            :
        ] = arr

        edge_size = min(
            EDGE_SAMPLE_PX,
            height,
        )

        top_strip = arr[
            :edge_size,
            :,
            :
        ]

        bottom_strip = arr[
            -edge_size:,
            :,
            :
        ]

        if pad_top > 0:

            fill = build_edge_fill(
                top_strip,
                pad_top,
                axis="v",
            )

            fill = add_matched_noise(
                fill,
                estimate_local_noise(
                    top_strip
                ),
                seed=1,
            )

            canvas[
                :pad_top,
                :,
                :
            ] = fill

        if pad_bottom > 0:

            fill = build_edge_fill(
                bottom_strip,
                pad_bottom,
                axis="v",
            )

            fill = add_matched_noise(
                fill,
                estimate_local_noise(
                    bottom_strip
                ),
                seed=2,
            )

            canvas[
                pad_top + height:,
                :,
                :
            ] = fill

        seams = []

        if pad_top > 0:

            seams.append(
                pad_top
            )

        if pad_bottom > 0:

            seams.append(
                pad_top
                + height
            )

        vertical_padding = True

        transformation = (
            f"PAD_TOP_BOTTOM | "
            f"top={pad_top}px | "
            f"bottom={pad_bottom}px"
        )

    # ========================================================
    # IMAGE TOO TALL / NARROW
    #
    # Add LEFT + RIGHT
    # ========================================================

    else:

        new_height = height

        new_width = round(
            height
            / TARGET_RATIO
        )

        pad_total = (
            new_width
            - width
        )

        pad_left = (
            pad_total
            // 2
        )

        pad_right = (
            pad_total
            - pad_left
        )

        canvas = np.zeros(
            (
                new_height,
                new_width,
                3,
            ),
            dtype=np.float32,
        )

        # Original image remains untouched.
        canvas[
            :,
            pad_left:
            pad_left + width,
            :
        ] = arr

        edge_size = min(
            EDGE_SAMPLE_PX,
            width,
        )

        left_strip = arr[
            :,
            :edge_size,
            :
        ]

        right_strip = arr[
            :,
            -edge_size:,
            :
        ]

        if pad_left > 0:

            fill = build_edge_fill(
                left_strip,
                pad_left,
                axis="h",
            )

            fill = add_matched_noise(
                fill,
                estimate_local_noise(
                    left_strip
                ),
                seed=3,
            )

            canvas[
                :,
                :pad_left,
                :
            ] = fill

        if pad_right > 0:

            fill = build_edge_fill(
                right_strip,
                pad_right,
                axis="h",
            )

            fill = add_matched_noise(
                fill,
                estimate_local_noise(
                    right_strip
                ),
                seed=4,
            )

            canvas[
                :,
                pad_left + width:,
                :
            ] = fill

        seams = []

        if pad_left > 0:

            seams.append(
                pad_left
            )

        if pad_right > 0:

            seams.append(
                pad_left
                + width
            )

        vertical_padding = False

        transformation = (
            f"PAD_LEFT_RIGHT | "
            f"left={pad_left}px | "
            f"right={pad_right}px"
        )

    # ========================================================
    # CONVERT BACK TO PIL
    # ========================================================

    output_image = (
        Image.fromarray(
            np.clip(
                canvas,
                0,
                255,
            ).astype(
                np.uint8
            )
        )
    )

    # ========================================================
    # SEAM BLENDING
    #
    # Safe because images touching boundaries were already
    # excluded before reaching this function.
    # ========================================================

    for seam in seams:

        if vertical_padding:

            box = (
                0,
                max(
                    0,
                    seam
                    - SEAM_BLUR_BAND,
                ),
                new_width,
                min(
                    new_height,
                    seam
                    + SEAM_BLUR_BAND,
                ),
            )

        else:

            box = (
                max(
                    0,
                    seam
                    - SEAM_BLUR_BAND,
                ),
                0,
                min(
                    new_width,
                    seam
                    + SEAM_BLUR_BAND,
                ),
                new_height,
            )

        crop = (
            output_image.crop(
                box
            )
        )

        crop = crop.filter(
            ImageFilter.GaussianBlur(
                radius=SEAM_BLUR_RADIUS
            )
        )

        output_image.paste(
            crop,
            box,
        )

    return (
        output_image,
        transformation,
    )


# ============================================================
# SOURCE URL NORMALIZATION
# ============================================================


def normalize_source_url(url):
    """
    Convert Google Drive share URLs into direct-download URLs.

    Other image URLs remain unchanged.
    """

    url = str(
        url
    ).strip()

    # --------------------------------------------------------
    # Example:
    #
    # drive.google.com/file/d/FILE_ID/view
    # --------------------------------------------------------

    drive_match = re.search(
        r"drive\.google\.com/file/d/([^/?]+)",
        url,
        re.IGNORECASE,
    )

    if drive_match:

        file_id = (
            drive_match.group(1)
        )

        return (
            "https://drive.usercontent.google.com/"
            f"download?id={file_id}&export=download"
        )

    # --------------------------------------------------------
    # Example:
    #
    # drive.google.com/open?id=FILE_ID
    # --------------------------------------------------------

    if "drive.google.com" in url:

        id_match = re.search(
            r"[?&]id=([^&]+)",
            url,
            re.IGNORECASE,
        )

        if id_match:

            file_id = (
                id_match.group(1)
            )

            return (
                "https://drive.usercontent.google.com/"
                f"download?id={file_id}&export=download"
            )

    return url


# ============================================================
# IMAGE DOWNLOAD
# ============================================================


def download_image(url):
    """
    Download image with retries, redirect handling and size limit.
    """

    url = normalize_source_url(
        url
    )

    headers = {
        "User-Agent": (
            "Mozilla/5.0 "
            "(Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 "
            "(KHTML, like Gecko) "
            "Chrome/151.0 Safari/537.36"
        ),
        "Accept": (
            "image/avif,"
            "image/webp,"
            "image/apng,"
            "image/*,"
            "*/*;q=0.8"
        ),
        "Accept-Language": (
            "en-US,en;q=0.9"
        ),
    }

    max_bytes = (
        MAX_DOWNLOAD_MB
        * 1024
        * 1024
    )

    last_error = None

    for attempt in range(
        1,
        DOWNLOAD_RETRIES + 1,
    ):

        try:

            with requests.get(
                url,
                headers=headers,
                timeout=(
                    CONNECT_TIMEOUT,
                    READ_TIMEOUT,
                ),
                stream=True,
                allow_redirects=True,
            ) as response:

                # Temporary failures can be retried.
                if response.status_code in (
                    429,
                    500,
                    502,
                    503,
                    504,
                ):

                    raise (
                        requests
                        .exceptions
                        .HTTPError(
                            f"HTTP "
                            f"{response.status_code}",
                            response=response,
                        )
                    )

                response.raise_for_status()

                content_type = (
                    response.headers
                    .get(
                        "Content-Type",
                        "",
                    )
                    .lower()
                )

                if (
                    content_type
                    and "image/" not in content_type
                    and "application/octet-stream"
                    not in content_type
                ):

                    raise ValueError(
                        "URL did not return an image. "
                        f"Content-Type={content_type}"
                    )

                content_length = (
                    response.headers.get(
                        "Content-Length"
                    )
                )

                if content_length:

                    try:

                        content_length_int = int(
                            content_length
                        )

                    except (
                        ValueError,
                        TypeError,
                    ):

                        content_length_int = None

                    if (
                        content_length_int
                        is not None
                        and content_length_int
                        > max_bytes
                    ):

                        raise ValueError(
                            f"Image is larger than "
                            f"{MAX_DOWNLOAD_MB} MB"
                        )

                buffer = BytesIO()

                downloaded = 0

                for chunk in (
                    response.iter_content(
                        chunk_size=64
                        * 1024
                    )
                ):

                    if not chunk:
                        continue

                    downloaded += len(
                        chunk
                    )

                    if downloaded > max_bytes:

                        raise ValueError(
                            f"Image is larger than "
                            f"{MAX_DOWNLOAD_MB} MB"
                        )

                    buffer.write(
                        chunk
                    )

            buffer.seek(
                0
            )

            image = Image.open(
                buffer
            )

            # Decode while buffer still exists.
            image.load()

            return image

        except (
            requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.HTTPError,
        ) as error:

            last_error = error

            status_code = None

            if (
                isinstance(
                    error,
                    requests.exceptions.HTTPError,
                )
                and error.response
                is not None
            ):

                status_code = (
                    error
                    .response
                    .status_code
                )

            # Don't retry permanent HTTP errors such as 403/404.
            if (
                status_code
                is not None
                and status_code
                not in (
                    429,
                    500,
                    502,
                    503,
                    504,
                )
            ):

                raise

            if attempt < DOWNLOAD_RETRIES:

                time.sleep(
                    attempt
                )

    if last_error is not None:

        raise last_error

    raise RuntimeError(
        "Image download failed."
    )


# ============================================================
# GENERAL HELPERS
# ============================================================


def safe_filename(value):
    """
    Convert Parent Group Id into filesystem-safe text.
    """

    value = str(
        value
    ).strip()

    value = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        value,
    )

    value = re.sub(
        r"\s+",
        "_",
        value,
    )

    value = value.strip(
        "._"
    )

    return (
        value
        or "unknown"
    )


def get_image_source(cell):
    """
    Read URL from either normal Excel cell or hyperlink.
    """

    if cell.hyperlink:

        target = (
            cell.hyperlink.target
        )

        if target:

            return str(
                target
            ).strip()

    if cell.value is None:

        return None

    value = str(
        cell.value
    ).strip()

    return (
        value
        or None
    )


# ============================================================
# EXCEL HEADER DETECTION
# ============================================================


def find_header_column(
    worksheet,
    expected_header,
):
    """
    Find header in row 1, case-insensitive.
    """

    expected = (
        expected_header
        .strip()
        .lower()
    )

    for cell in worksheet[1]:

        if cell.value is None:
            continue

        actual = (
            str(
                cell.value
            )
            .strip()
            .lower()
        )

        if actual == expected:

            return cell.column

    return None


def find_image_columns(
    worksheet,
):
    """
    Automatically detect:

        Image Url 1
        Image Url 2
        ...
        Image Url 8
    """

    pattern = re.compile(
        r"^image\s*url\s*(\d+)$",
        re.IGNORECASE,
    )

    image_columns = []

    for cell in worksheet[1]:

        if cell.value is None:
            continue

        header = str(
            cell.value
        ).strip()

        match = pattern.match(
            header
        )

        if not match:
            continue

        position = int(
            match.group(1)
        )

        if not (
            1
            <= position
            <= MAX_IMAGES_PER_SKU
        ):

            continue

        image_columns.append(
            {
                "position": position,
                "column": cell.column,
                "header": header,
            }
        )

    image_columns.sort(
        key=lambda item: (
            item["position"]
        )
    )

    return image_columns


# ============================================================
# PROCESS ONE IMAGE
# ============================================================


def process_image(
    source_url,
    parent_group_id,
    image_position,
    output_dir,
):
    """
    Process one image.

    Decision flow:

        Download
            ↓
        Minimum resolution?
            ↓
        Product touches any edge?
           YES → SKIP
           NO
            ↓
        Ratio correct?
           YES → save unchanged
           NO  → transform + save
    """

    result = {
        "parent_group_id": (
            str(
                parent_group_id
            )
        ),
        "image_position": (
            image_position
        ),
        "source_url": (
            source_url
        ),
        "processing_status": "",
        "original_dimensions": "",
        "original_w_h_ratio": "",
        "edge_warning": "",
        "edge_detection_details": "",
        "output_dimensions": "",
        "output_w_h_ratio": "",
        "transformation": "",
        "output_file": "",
        "processing_error": "",
    }

    try:

        # ====================================================
        # DOWNLOAD
        # ====================================================

        image = download_image(
            source_url
        )

        image = normalize_image(
            image
        )

        original_width, original_height = (
            image.size
        )

        result[
            "original_dimensions"
        ] = (
            f"{original_width}x"
            f"{original_height}"
        )

        result[
            "original_w_h_ratio"
        ] = round(
            original_width
            / original_height,
            4,
        )

        # ====================================================
        # MINIMUM RESOLUTION
        # ====================================================

        if (
            original_width < MIN_W
            or original_height < MIN_H
        ):

            result[
                "processing_status"
            ] = (
                "BLOCKED_MIN_RESOLUTION"
            )

            result[
                "processing_error"
            ] = (
                f"Image "
                f"{original_width}x"
                f"{original_height} "
                f"is below minimum "
                f"{MIN_W}x{MIN_H}"
            )

            return result

        # ====================================================
        # SELLER OPS RULE
        #
        # Product touches ANY boundary:
        #
        # DO NOT TRANSFORM
        # DO NOT RE-ENCODE
        # DO NOT CREATE OUTPUT IMAGE
        # ====================================================

        (
            edge_warnings,
            fractions,
        ) = detect_product_touching_edges(
            image
        )

        result[
            "edge_detection_details"
        ] = (
            " | ".join(
                f"{name}="
                f"{fraction:.1%}"
                for name, fraction
                in fractions.items()
            )
        )

        if edge_warnings:

            result[
                "processing_status"
            ] = (
                "SKIPPED_PRODUCT_TOUCHES_EDGE"
            )

            result[
                "edge_warning"
            ] = (
                "; ".join(
                    edge_warnings
                )
            )

            result[
                "transformation"
            ] = (
                "SKIPPED"
            )

            return result

        # ====================================================
        # SAFE TO TRANSFORM
        # ====================================================

        (
            transformed,
            transformation,
        ) = pad_to_target_ratio(
            image
        )

        output_width, output_height = (
            transformed.size
        )

        result[
            "output_dimensions"
        ] = (
            f"{output_width}x"
            f"{output_height}"
        )

        result[
            "output_w_h_ratio"
        ] = round(
            output_width
            / output_height,
            4,
        )

        result[
            "transformation"
        ] = (
            transformation
        )

        # ====================================================
        # OUTPUT FILE
        #
        # No leading zero:
        #
        # SKU_1.jpg
        # SKU_2.jpg
        # ====================================================

        filename = (
            f"{safe_filename(parent_group_id)}_"
            f"{image_position}.jpg"
        )

        output_path = (
            os.path.join(
                output_dir,
                filename,
            )
        )

        transformed.save(
            output_path,
            format="JPEG",
            quality=95,
            optimize=True,
            subsampling=0,
        )

        result[
            "output_file"
        ] = filename

        # ====================================================
        # SUCCESS
        # ====================================================

        if transformation == "NONE":

            result[
                "processing_status"
            ] = (
                "SUCCESS_NO_CHANGE"
            )

        else:

            result[
                "processing_status"
            ] = (
                "SUCCESS_TRANSFORMED"
            )

        return result

    # ========================================================
    # ERRORS
    # ========================================================

    except requests.exceptions.Timeout:

        result[
            "processing_status"
        ] = (
            "ERROR_TIMEOUT"
        )

        result[
            "processing_error"
        ] = (
            "Image download timed out"
        )

        return result

    except requests.exceptions.HTTPError as error:

        status_code = ""

        if (
            error.response
            is not None
        ):

            status_code = (
                error
                .response
                .status_code
            )

        result[
            "processing_status"
        ] = (
            "ERROR_HTTP"
        )

        result[
            "processing_error"
        ] = (
            f"HTTP "
            f"{status_code}: "
            f"{error}"
        )

        return result

    except requests.exceptions.ConnectionError as error:

        result[
            "processing_status"
        ] = (
            "ERROR_CONNECTION"
        )

        result[
            "processing_error"
        ] = (
            str(
                error
            )
        )

        return result

    except Exception as error:

        result[
            "processing_status"
        ] = (
            "ERROR"
        )

        result[
            "processing_error"
        ] = (
            str(
                error
            )
        )

        return result


# ============================================================
# PROCESSING RESULTS SHEET
# ============================================================


RESULT_HEADERS = [
    "Parent Group Id",
    "Image Position",
    "Source Column",
    "Source URL",
    "Processing Status",
    "Original Dimensions",
    "Original W:H Ratio",
    "Edge Warning",
    "Edge Detection Details",
    "Output Dimensions",
    "Output W:H Ratio",
    "Transformation",
    "Output File",
    "Processing Error",
]


def prepare_results_sheet(
    workbook,
):
    """
    Recreate Processing Results sheet.
    """

    if (
        RESULT_SHEET_NAME
        in workbook.sheetnames
    ):

        workbook.remove(
            workbook[
                RESULT_SHEET_NAME
            ]
        )

    worksheet = (
        workbook.create_sheet(
            RESULT_SHEET_NAME
        )
    )

    for (
        column_index,
        header,
    ) in enumerate(
        RESULT_HEADERS,
        start=1,
    ):

        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        cell.alignment = (
            Alignment(
                vertical="center",
            )
        )

    worksheet.freeze_panes = (
        "A2"
    )

    return worksheet


def write_result_row(
    worksheet,
    source_column,
    result,
):
    """
    Write one image processing result.
    """

    worksheet.append(
        [
            result[
                "parent_group_id"
            ],
            result[
                "image_position"
            ],
            source_column,
            result[
                "source_url"
            ],
            result[
                "processing_status"
            ],
            result[
                "original_dimensions"
            ],
            result[
                "original_w_h_ratio"
            ],
            result[
                "edge_warning"
            ],
            result[
                "edge_detection_details"
            ],
            result[
                "output_dimensions"
            ],
            result[
                "output_w_h_ratio"
            ],
            result[
                "transformation"
            ],
            result[
                "output_file"
            ],
            result[
                "processing_error"
            ],
        ]
    )


def format_results_sheet(
    worksheet,
):
    """
    Format Processing Results sheet.
    """

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    widths = {
        1: 30,
        2: 15,
        3: 18,
        4: 70,
        5: 36,
        6: 22,
        7: 20,
        8: 45,
        9: 55,
        10: 22,
        11: 20,
        12: 42,
        13: 42,
        14: 65,
    }

    for (
        column_index,
        width,
    ) in widths.items():

        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = width

    success_fill = (
        PatternFill(
            fill_type="solid",
            fgColor="E2F0D9",
        )
    )

    warning_fill = (
        PatternFill(
            fill_type="solid",
            fgColor="FCE5CD",
        )
    )

    error_fill = (
        PatternFill(
            fill_type="solid",
            fgColor="F4CCCC",
        )
    )

    for row_index in range(
        2,
        worksheet.max_row
        + 1,
    ):

        status_cell = (
            worksheet.cell(
                row=row_index,
                column=5,
            )
        )

        status = str(
            status_cell.value
            or ""
        )

        if status.startswith(
            "SUCCESS"
        ):

            status_cell.fill = (
                success_fill
            )

        elif (
            status.startswith(
                "BLOCKED"
            )
            or status.startswith(
                "SKIPPED"
            )
        ):

            status_cell.fill = (
                warning_fill
            )

        elif status.startswith(
            "ERROR"
        ):

            status_cell.fill = (
                error_fill
            )


# ============================================================
# EXCEL PROCESSOR
# ============================================================


def process_excel(
    input_path,
    output_dir,
    report_path,
    sku_column=DEFAULT_SKU_COLUMN,
    sheet_name=None,
    progress_callback=None,
):
    """
    Process complete catalogue workbook.
    """

    workbook = load_workbook(
        input_path
    )

    # ========================================================
    # INPUT SHEET
    # ========================================================

    if sheet_name:

        if (
            sheet_name
            not in workbook.sheetnames
        ):

            raise ValueError(
                f"Sheet "
                f"'{sheet_name}' "
                f"not found. "
                f"Available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = (
            workbook[
                sheet_name
            ]
        )

    else:

        possible_sheets = [
            sheet
            for sheet
            in workbook.worksheets
            if (
                sheet.title
                != RESULT_SHEET_NAME
            )
        ]

        if not possible_sheets:

            raise ValueError(
                "No input worksheet found."
            )

        worksheet = (
            possible_sheets[0]
        )

    input_sheet_title = (
        worksheet.title
    )

    # ========================================================
    # FIND SKU COLUMN
    # ========================================================

    sku_column_index = (
        find_header_column(
            worksheet,
            sku_column,
        )
    )

    if (
        sku_column_index
        is None
    ):

        headers = [
            str(
                cell.value
            )
            for cell
            in worksheet[1]
            if (
                cell.value
                is not None
            )
        ]

        raise ValueError(
            f"Could not find "
            f"'{sku_column}'. "
            f"Available headers: "
            f"{headers}"
        )

    # ========================================================
    # FIND IMAGE URL COLUMNS
    # ========================================================

    image_columns = (
        find_image_columns(
            worksheet
        )
    )

    if not image_columns:

        raise ValueError(
            "No image columns found. "
            "Expected "
            "'Image Url 1', "
            "'Image Url 2', etc."
        )

    # ========================================================
    # OUTPUT
    # ========================================================

    os.makedirs(
        output_dir,
        exist_ok=True,
    )

    results_sheet = (
        prepare_results_sheet(
            workbook
        )
    )

    # ========================================================
    # COUNT IMAGES
    # ========================================================

    total_images = 0

    for row_index in range(
        2,
        worksheet.max_row
        + 1,
    ):

        parent_group_id = (
            worksheet.cell(
                row=row_index,
                column=sku_column_index,
            ).value
        )

        if (
            parent_group_id
            is None
            or not str(
                parent_group_id
            ).strip()
        ):

            continue

        for item in image_columns:

            image_cell = (
                worksheet.cell(
                    row=row_index,
                    column=item[
                        "column"
                    ],
                )
            )

            if get_image_source(
                image_cell
            ):

                total_images += 1

    if total_images == 0:

        raise ValueError(
            "No image URLs were found."
        )

    print()

    print(
        f"Sheet: "
        f"{input_sheet_title}"
    )

    print(
        f"Images to process: "
        f"{total_images}"
    )

    print(
        f"Minimum resolution: "
        f"{MIN_W}x{MIN_H}"
    )

    print(
        f"Target W:H: "
        f"{TARGET_W_H_RATIO:.6f}"
    )

    print(
        "Edge rule: "
        "skip image if product "
        "touches any boundary"
    )

    print()

    # ========================================================
    # PROCESS
    # ========================================================

    status_counts = (
        Counter()
    )

    processed_count = 0

    for row_index in range(
        2,
        worksheet.max_row
        + 1,
    ):

        parent_group_id = (
            worksheet.cell(
                row=row_index,
                column=sku_column_index,
            ).value
        )

        if (
            parent_group_id
            is None
            or not str(
                parent_group_id
            ).strip()
        ):

            continue

        parent_group_id = (
            str(
                parent_group_id
            ).strip()
        )

        for item in image_columns:

            image_cell = (
                worksheet.cell(
                    row=row_index,
                    column=item[
                        "column"
                    ],
                )
            )

            source_url = (
                get_image_source(
                    image_cell
                )
            )

            if not source_url:

                continue

            # ------------------------------------------------
            # STREAMLIT PROGRESS - STARTING
            # ------------------------------------------------

            if progress_callback:

                progress_callback(
                    processed_count,
                    total_images,
                    parent_group_id,
                    item[
                        "position"
                    ],
                    "PROCESSING",
                )

            # ------------------------------------------------
            # PROCESS IMAGE
            # ------------------------------------------------

            result = process_image(
                source_url=source_url,
                parent_group_id=(
                    parent_group_id
                ),
                image_position=item[
                    "position"
                ],
                output_dir=(
                    output_dir
                ),
            )

            processed_count += 1

            status = result[
                "processing_status"
            ]

            status_counts[
                status
            ] += 1

            # ------------------------------------------------
            # WRITE REPORT
            # ------------------------------------------------

            write_result_row(
                worksheet=(
                    results_sheet
                ),
                source_column=item[
                    "header"
                ],
                result=result,
            )

            # ------------------------------------------------
            # STREAMLIT PROGRESS - COMPLETE
            # ------------------------------------------------

            if progress_callback:

                progress_callback(
                    processed_count,
                    total_images,
                    parent_group_id,
                    item[
                        "position"
                    ],
                    status,
                )

            print(
                f"["
                f"{processed_count}"
                f"/"
                f"{total_images}"
                f"] "
                f"{parent_group_id} "
                f"| Image "
                f"{item['position']} "
                f"| {status}"
            )

            if result[
                "edge_warning"
            ]:

                print(
                    f"    -> "
                    f"{result['edge_warning']}"
                )

    # ========================================================
    # SAVE REPORT
    # ========================================================

    format_results_sheet(
        results_sheet
    )

    workbook.save(
        report_path
    )

    workbook.close()

    # ========================================================
    # SUMMARY
    # ========================================================

    successful = sum(
        count
        for status, count
        in status_counts.items()
        if status.startswith(
            "SUCCESS"
        )
    )

    transformed = (
        status_counts.get(
            "SUCCESS_TRANSFORMED",
            0,
        )
    )

    unchanged = (
        status_counts.get(
            "SUCCESS_NO_CHANGE",
            0,
        )
    )

    skipped = (
        status_counts.get(
            "SKIPPED_PRODUCT_TOUCHES_EDGE",
            0,
        )
    )

    blocked = (
        status_counts.get(
            "BLOCKED_MIN_RESOLUTION",
            0,
        )
    )

    errors = sum(
        count
        for status, count
        in status_counts.items()
        if status.startswith(
            "ERROR"
        )
    )

    return {
        "total": (
            processed_count
        ),
        "successful": (
            successful
        ),
        "transformed": (
            transformed
        ),
        "unchanged": (
            unchanged
        ),
        "skipped": (
            skipped
        ),
        "blocked": (
            blocked
        ),
        "errors": (
            errors
        ),
        "statuses": dict(
            status_counts
        ),
        "image_columns": [
            item[
                "header"
            ]
            for item
            in image_columns
        ],
        "input_sheet": (
            input_sheet_title
        ),
    }


# ============================================================
# COMMAND LINE
# ============================================================


def main():

    parser = argparse.ArgumentParser(
        description=(
            "Validate and transform "
            "product images from "
            "Parent Group Id + "
            "Image Url 1..8."
        )
    )

    parser.add_argument(
        "input",
        help=(
            "Input .xlsx catalogue"
        ),
    )

    parser.add_argument(
        "output",
        help=(
            "Directory for "
            "transformed images"
        ),
    )

    parser.add_argument(
        "--sheet",
        default=None,
        help=(
            "Optional Excel "
            "sheet name"
        ),
    )

    parser.add_argument(
        "--sku-column",
        default=(
            DEFAULT_SKU_COLUMN
        ),
        help=(
            "Default: "
            "Parent Group Id"
        ),
    )

    parser.add_argument(
        "--report",
        default=None,
        help=(
            "Optional output "
            "report path"
        ),
    )

    args = parser.parse_args()

    input_path = (
        os.path.abspath(
            args.input
        )
    )

    output_dir = (
        os.path.abspath(
            args.output
        )
    )

    if not os.path.isfile(
        input_path
    ):

        raise FileNotFoundError(
            f"Input file not found: "
            f"{input_path}"
        )

    if not (
        input_path
        .lower()
        .endswith(
            ".xlsx"
        )
    ):

        raise ValueError(
            "Input must be "
            "an .xlsx file."
        )

    os.makedirs(
        output_dir,
        exist_ok=True,
    )

    if args.report:

        report_path = (
            os.path.abspath(
                args.report
            )
        )

    else:

        report_path = (
            os.path.join(
                output_dir,
                "processing_report.xlsx",
            )
        )

    summary = (
        process_excel(
            input_path=(
                input_path
            ),
            output_dir=(
                output_dir
            ),
            report_path=(
                report_path
            ),
            sku_column=(
                args.sku_column
            ),
            sheet_name=(
                args.sheet
            ),
        )
    )

    print()

    print(
        "=" * 60
    )

    print(
        "PROCESSING COMPLETE"
    )

    print(
        "=" * 60
    )

    print(
        f"Total: "
        f"{summary['total']}"
    )

    print(
        f"Successful: "
        f"{summary['successful']}"
    )

    print(
        f"Transformed: "
        f"{summary['transformed']}"
    )

    print(
        f"Unchanged: "
        f"{summary['unchanged']}"
    )

    print(
        "Skipped - product "
        "touches edge: "
        f"{summary['skipped']}"
    )

    print(
        "Blocked - minimum "
        "resolution: "
        f"{summary['blocked']}"
    )

    print(
        f"Errors: "
        f"{summary['errors']}"
    )

    print()

    print(
        f"Report: "
        f"{report_path}"
    )


if __name__ == "__main__":

    try:

        main()

    except KeyboardInterrupt:

        print(
            "\nProcessing cancelled."
        )

        sys.exit(
            1
        )

    except Exception as error:

        print()

        print(
            f"ERROR: "
            f"{error}"
        )

        sys.exit(
            1
        )